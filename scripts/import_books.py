#!/usr/bin/env python3
"""
import_books.py
--------------
1. Reads 500_livres_isbn.xlsx (columns: N°, Titre, Auteur, ISBN-13)
2. Queries Google Books API per ISBN to verify title/author
3. Adds a "véracité" column: "vrai" if match, "faux" otherwise
4. Saves the annotated Excel file back
5. For every "vrai" row, inserts author(s) and book into Supabase

Usage:
    python scripts/import_books.py

Required environment variables (or edit SUPABASE_* constants below):
    SUPABASE_URL
    SUPABASE_SERVICE_ROLE_KEY   ← from Supabase > Project Settings > API
"""

import os
import re
import time
import unicodedata
import requests
import openpyxl
from supabase import create_client, Client

# ──────────────────────────────────────────────
# CONFIG — edit or set as env vars
# ──────────────────────────────────────────────
EXCEL_PATH = os.path.join(
    os.path.dirname(__file__), "..", "..", "500_livres_isbn.xlsx"
)

SUPABASE_URL = os.getenv("SUPABASE_URL", "https://jtgdtydqekakkrvoekqs.supabase.co")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")  # REQUIRED

GOOGLE_BOOKS_API = "https://www.googleapis.com/books/v1/volumes"
GOOGLE_API_KEY   = os.getenv("GOOGLE_BOOKS_API_KEY", "")   # optional but avoids rate limits

# How similar do titles/authors need to be? (0.0–1.0)
MATCH_THRESHOLD = 0.6

# Seconds between API calls (be polite)
API_DELAY = 0.4


# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────

def normalize(text: str) -> str:
    """Lowercase, strip accents, remove punctuation/extra spaces."""
    if not text:
        return ""
    text = str(text)
    # NFD decompose → remove combining marks (accents)
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.lower()
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def token_overlap(a: str, b: str) -> float:
    """Jaccard-style token overlap between two normalized strings."""
    sa = set(normalize(a).split())
    sb = set(normalize(b).split())
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / len(sa | sb)


def split_authors(raw: str) -> list[str]:
    """Split an author string on commas and semicolons, strip whitespace."""
    parts = re.split(r"[,;]+", raw or "")
    return [p.strip() for p in parts if p.strip()]


def to_hd_cover(url: str | None) -> str | None:
    """Convert Google Books cover URL to high-res version."""
    if not url:
        return None
    url = url.replace("http://", "https://")
    url = url.replace("&edge=curl", "")
    url = re.sub(r"zoom=\d", "zoom=0", url)
    return url + "&fife=w600"


def google_books_search(isbn: str) -> dict | None:
    """
    Query Google Books by ISBN-13.
    Returns the first volume item dict, or None.
    """
    params: dict = {"q": f"isbn:{isbn}", "maxResults": 1}
    if GOOGLE_API_KEY:
        params["key"] = GOOGLE_API_KEY

    try:
        resp = requests.get(GOOGLE_BOOKS_API, params=params, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        items = data.get("items")
        if items:
            return items[0]
    except Exception as e:
        print(f"  [WARN] Google Books error for ISBN {isbn}: {e}")
    return None


def google_books_search_by_title_author(title: str, author: str) -> dict | None:
    """Fallback: search by title + first author."""
    first_author = split_authors(author)[0] if author else ""
    query = f"intitle:{title}"
    if first_author:
        query += f"+inauthor:{first_author}"
    params: dict = {"q": query, "maxResults": 1}
    if GOOGLE_API_KEY:
        params["key"] = GOOGLE_API_KEY

    try:
        resp = requests.get(GOOGLE_BOOKS_API, params=params, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        items = data.get("items")
        if items:
            return items[0]
    except Exception as e:
        print(f"  [WARN] Fallback search error: {e}")
    return None


def verify_match(item: dict, expected_title: str, expected_author: str) -> bool:
    """Check if a Google Books item matches the expected title and author."""
    info = item.get("volumeInfo", {})
    api_title   = info.get("title", "")
    api_authors = info.get("authors", [])

    title_score = token_overlap(api_title, expected_title)

    author_score = 0.0
    for api_author in api_authors:
        for exp_author in split_authors(expected_author):
            s = token_overlap(api_author, exp_author)
            if s > author_score:
                author_score = s

    match = title_score >= MATCH_THRESHOLD and author_score >= MATCH_THRESHOLD
    return match


def extract_book_data(item: dict, isbn: str) -> dict:
    """Pull relevant fields from a Google Books volume item."""
    info = item.get("volumeInfo", {})
    image_links = info.get("imageLinks", {})
    raw_cover   = image_links.get("extraLarge") or image_links.get("large") or image_links.get("thumbnail")

    industry_ids = info.get("industryIdentifiers", [])
    isbn13 = isbn  # default to what we searched
    isbn10 = None
    for id_obj in industry_ids:
        if id_obj.get("type") == "ISBN_13":
            isbn13 = id_obj.get("identifier", isbn)
        elif id_obj.get("type") == "ISBN_10":
            isbn10 = id_obj.get("identifier")

    return {
        "title":        info.get("title", ""),
        "description":  info.get("description"),
        "published_at": info.get("publishedDate"),
        "page_count":   info.get("pageCount"),
        "cover_url":    to_hd_cover(raw_cover),
        "isbn_13":      isbn13,
        "isbn_10":      isbn10,
        "language":     info.get("language"),
        "api_authors":  info.get("authors", []),
        "categories":   info.get("categories", []),
    }


# ──────────────────────────────────────────────
# SUPABASE HELPERS
# ──────────────────────────────────────────────

def upsert_author(sb: Client, name: str) -> str | None:
    """
    Insert author if not exists (match on normalized name).
    Returns the author UUID.
    """
    if not name:
        return None

    # Check existing by name (case-insensitive)
    res = sb.table("authors").select("id").ilike("name", name).limit(1).execute()
    if res.data:
        return res.data[0]["id"]

    # Insert new
    ins = sb.table("authors").insert({"name": name}).execute()
    if ins.data:
        return ins.data[0]["id"]
    print(f"  [ERROR] Could not insert author: {name}")
    return None


def book_exists(sb: Client, title: str, author_id: str | None) -> bool:
    """Check if a book with the same normalized title and author already exists."""
    res = (
        sb.table("books")
        .select("id")
        .ilike("title", title)
        .eq("author_id", author_id)
        .limit(1)
        .execute()
    )
    return bool(res.data)


def insert_book(sb: Client, book_data: dict, author_ids: list[str]) -> str | None:
    """
    Insert book into the books table using its actual schema.
    Returns the book UUID.
    """
    primary_author_id = author_ids[0] if author_ids else None

    payload = {
        "title":       book_data["title"],
        "description": book_data["description"],
        "cover_url":   book_data["cover_url"],
        "author_id":   primary_author_id,
    }
    # Remove None values to let DB defaults apply
    payload = {k: v for k, v in payload.items() if v is not None}

    res = sb.table("books").insert(payload).execute()
    if not res.data:
        print(f"  [ERROR] Could not insert book: {book_data['title']}")
        return None

    return res.data[0]["id"]


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────

def main():
    if not SUPABASE_KEY:
        print("ERROR: SUPABASE_SERVICE_ROLE_KEY is not set.")
        print("Export it before running:")
        print("  export SUPABASE_SERVICE_ROLE_KEY='your-service-role-key'")
        return

    # Load Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Find or create "véracité" column
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")

    veracite_col = None
    for idx, h in enumerate(headers, start=1):
        if h and normalize(str(h)) == normalize("véracité"):
            veracite_col = idx
            break

    if veracite_col is None:
        veracite_col = len(headers) + 1
        ws.cell(row=1, column=veracite_col, value="véracité")
        print(f"Created 'véracité' column at column {veracite_col}")

    # Map column names → indices
    col_map = {normalize(str(h)): idx for idx, h in enumerate(headers, start=1) if h}
    print(f"Column map: {col_map}")

    # Identify required columns
    titre_col   = col_map.get("titre")
    auteur_col  = col_map.get("auteur")
    isbn_col    = col_map.get(normalize("ISBN-13")) or col_map.get("isbn 13") or col_map.get("isbn13")

    if not all([titre_col, auteur_col, isbn_col]):
        print(f"ERROR: Could not find required columns. Found: {list(col_map.keys())}")
        return

    # Init Supabase
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    print(f"Connected to Supabase: {SUPABASE_URL}\n")

    stats = {"total": 0, "vrai": 0, "faux": 0, "inserted": 0, "skipped": 0, "error": 0}

    for row_num in range(2, ws.max_row + 1):
        titre  = ws.cell(row=row_num, column=titre_col).value
        auteur = ws.cell(row=row_num, column=auteur_col).value
        isbn   = ws.cell(row=row_num, column=isbn_col).value

        if not isbn and not titre:
            continue  # skip empty rows

        stats["total"] += 1
        isbn_str = str(isbn).strip().replace("-", "") if isbn else ""
        titre_str  = str(titre).strip() if titre else ""
        auteur_str = str(auteur).strip() if auteur else ""

        print(f"[{row_num-1:03d}] {titre_str[:45]:<45} | ISBN: {isbn_str}")

        # ── Phase 1: search by ISBN
        item = None
        if isbn_str:
            item = google_books_search(isbn_str)
            time.sleep(API_DELAY)

        # ── Phase 2: fallback by title + author
        if item is None and titre_str:
            print(f"       → ISBN not found, trying title+author fallback")
            item = google_books_search_by_title_author(titre_str, auteur_str)
            time.sleep(API_DELAY)

        # ── Verify match
        if item is None:
            verdict = "faux"
            print(f"       → FAUX (no API result)")
        elif verify_match(item, titre_str, auteur_str):
            verdict = "vrai"
            print(f"       → VRAI ✓")
        else:
            api_info = item.get("volumeInfo", {})
            print(f"       → FAUX (API: '{api_info.get('title','')}' / {api_info.get('authors',[])})")
            verdict = "faux"

        ws.cell(row=row_num, column=veracite_col, value=verdict)

        if verdict == "vrai":
            stats["vrai"] += 1
        else:
            stats["faux"] += 1
            continue

        # ── Insert into Supabase
        book_data = extract_book_data(item, isbn_str)

        # Determine authors: prefer API authors, fall back to Excel
        raw_authors = book_data["api_authors"] if book_data["api_authors"] else split_authors(auteur_str)
        author_ids = []
        for author_name in raw_authors:
            aid = upsert_author(sb, author_name)
            if aid:
                author_ids.append(aid)

        primary_author_id = author_ids[0] if author_ids else None

        if book_exists(sb, book_data["title"], primary_author_id):
            print(f"       → Already in DB, skipping")
            stats["skipped"] += 1
            continue

        book_id = insert_book(sb, book_data, author_ids)
        if book_id:
            print(f"       → Inserted book {book_id}")
            stats["inserted"] += 1
        else:
            stats["error"] += 1

    # Save Excel with véracité column
    wb.save(EXCEL_PATH)
    print(f"\nSaved updated Excel to {EXCEL_PATH}")

    print("\n── SUMMARY ──────────────────────────────")
    print(f"  Processed : {stats['total']}")
    print(f"  Vrai      : {stats['vrai']}")
    print(f"  Faux      : {stats['faux']}")
    print(f"  Inserted  : {stats['inserted']}")
    print(f"  Skipped   : {stats['skipped']}  (already in DB)")
    print(f"  Errors    : {stats['error']}")
    print("─────────────────────────────────────────")


if __name__ == "__main__":
    main()
